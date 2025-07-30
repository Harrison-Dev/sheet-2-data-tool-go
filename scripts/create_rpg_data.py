#!/usr/bin/env python3
import openpyxl
from openpyxl import Workbook
import os

def create_characters_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Characters"
    
    # Headers
    headers = ["ID", "Name", "JobClass", "Level", "HP", "MP", "Strength", "Agility", "Intelligence", "Luck", "Weapon", "Armor"]
    for i, header in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=header)
    
    # Sample character data
    characters = [
        [1, "Aragorn", "Warrior", 25, 350, 50, 85, 70, 45, 60, "Legendary Sword", "Plate Mail"],
        [2, "Gandalf", "Wizard", 99, 200, 500, 30, 40, 95, 80, "Staff of Power", "Wizard Robe"],
        [3, "Legolas", "Archer", 20, 280, 80, 60, 95, 55, 75, "Elven Bow", "Leather Armor"],
        [4, "Gimli", "Dwarf Fighter", 18, 320, 30, 80, 50, 35, 55, "Battle Axe", "Chain Mail"],
        [5, "Frodo", "Hobbit", 5, 150, 40, 25, 85, 60, 90, "Sting", "Mithril Shirt"],
        [6, "Merlin", "Archmage", 75, 180, 600, 35, 45, 99, 85, "Ancient Staff", "Arcane Robes"],
        [7, "Robin Hood", "Ranger", 22, 260, 90, 55, 90, 65, 70, "Longbow", "Studded Leather"],
        [8, "Conan", "Barbarian", 30, 400, 20, 90, 65, 40, 50, "Broadsword", "Barbarian Leather"]
    ]
    
    for i, character in enumerate(characters, 2):
        for j, value in enumerate(character, 1):
            ws.cell(row=i, column=j, value=value)
    
    wb.save("characters.xlsx")
    print("Created characters.xlsx")

def create_items_excel():
    wb = Workbook()
    
    # Weapons sheet
    ws_weapons = wb.active
    ws_weapons.title = "Weapons"
    
    weapon_headers = ["ID", "Name", "WeaponType", "Damage", "Durability", "Weight", "Value", "Rarity", "Requirements"]
    for i, header in enumerate(weapon_headers, 1):
        ws_weapons.cell(row=1, column=i, value=header)
    
    weapons = [
        [1, "Iron Sword", "Sword", 25, 100, 3.5, 150, "Common", "Str 15"],
        [2, "Steel Bow", "Bow", 30, 80, 2.0, 200, "Common", "Agi 20"],
        [3, "Flame Staff", "Staff", 40, 60, 1.5, 500, "Rare", "Int 25"],
        [4, "Mythril Blade", "Sword", 55, 150, 3.0, 1200, "Epic", "Str 30"],
        [5, "Shadow Dagger", "Dagger", 35, 90, 1.0, 800, "Rare", "Agi 25"],
        [6, "Warhammer", "Hammer", 60, 120, 8.0, 300, "Uncommon", "Str 35"],
        [7, "Crystal Wand", "Wand", 45, 70, 0.5, 600, "Rare", "Int 20"],
        [8, "Dragon Slayer", "Greatsword", 80, 200, 6.0, 5000, "Legendary", "Str 40"]
    ]
    
    for i, weapon in enumerate(weapons, 2):
        for j, value in enumerate(weapon, 1):
            ws_weapons.cell(row=i, column=j, value=value)
    
    # Armor sheet
    ws_armor = wb.create_sheet("Armor")
    armor_headers = ["ID", "Name", "ArmorType", "Defense", "Weight", "Value", "Rarity", "Special_Effect"]
    for i, header in enumerate(armor_headers, 1):
        ws_armor.cell(row=1, column=i, value=header)
    
    armors = [
        [1, "Leather Vest", "Light", 15, 2.0, 50, "Common", "None"],
        [2, "Chain Mail", "Medium", 25, 15.0, 200, "Common", "None"],
        [3, "Plate Armor", "Heavy", 40, 35.0, 800, "Uncommon", "None"],
        [4, "Elven Cloak", "Light", 20, 1.0, 500, "Rare", "+10 Stealth"],
        [5, "Dragon Scale", "Heavy", 60, 25.0, 3000, "Epic", "Fire Resist"],
        [6, "Mage Robes", "Cloth", 10, 1.5, 300, "Uncommon", "+20 MP"],
        [7, "Shadow Armor", "Medium", 35, 8.0, 1500, "Rare", "Invisibility"],
        [8, "Holy Plate", "Heavy", 50, 30.0, 2500, "Epic", "Undead Protection"]
    ]
    
    for i, armor in enumerate(armors, 2):
        for j, value in enumerate(armor, 1):
            ws_armor.cell(row=i, column=j, value=value)
    
    wb.save("items.xlsx")
    print("Created items.xlsx")

def create_skills_excel():
    wb = Workbook()
    
    # Magic Skills
    ws_magic = wb.active
    ws_magic.title = "Magic_Skills"
    
    magic_headers = ["ID", "Name", "Element", "MP_Cost", "Damage", "Range", "Cast_Time", "Level_Required", "Description"]
    for i, header in enumerate(magic_headers, 1):
        ws_magic.cell(row=1, column=i, value=header)
    
    magic_skills = [
        [1, "Fireball", "Fire", 15, 35, "Medium", 2.0, 5, "Launches a ball of fire"],
        [2, "Ice Shard", "Ice", 12, 30, "Long", 1.5, 3, "Shoots sharp ice projectile"],
        [3, "Lightning Bolt", "Lightning", 20, 45, "Long", 1.0, 8, "Strikes with lightning"],
        [4, "Heal", "Light", 10, 0, "Touch", 3.0, 1, "Restores HP to target"],
        [5, "Shield", "Arcane", 8, 0, "Self", 2.0, 2, "Creates protective barrier"],
        [6, "Meteor", "Fire", 50, 100, "Area", 5.0, 20, "Summons falling meteor"],
        [7, "Blizzard", "Ice", 40, 80, "Area", 4.0, 18, "Creates ice storm"],
        [8, "Teleport", "Arcane", 25, 0, "Anywhere", 1.0, 15, "Instantly move to location"]
    ]
    
    for i, skill in enumerate(magic_skills, 2):
        for j, value in enumerate(skill, 1):
            ws_magic.cell(row=i, column=j, value=value)
    
    # Combat Skills
    ws_combat = wb.create_sheet("Combat_Skills")
    combat_headers = ["ID", "Name", "SkillType", "Stamina_Cost", "Damage_Multiplier", "Accuracy", "Level_Required", "Weapon_Type", "Description"]
    for i, header in enumerate(combat_headers, 1):
        ws_combat.cell(row=1, column=i, value=header)
    
    combat_skills = [
        [1, "Power Strike", "Attack", 10, 1.5, 90, 3, "Sword", "Powerful sword attack"],
        [2, "Precise Shot", "Ranged", 8, 1.3, 95, 5, "Bow", "Accurate arrow shot"],
        [3, "Whirlwind", "AOE", 20, 1.2, 85, 10, "Sword", "Spinning attack hits all"],
        [4, "Backstab", "Stealth", 15, 2.0, 80, 7, "Dagger", "Critical hit from behind"],
        [5, "Shield Bash", "Stun", 12, 0.8, 95, 4, "Shield", "Stuns target briefly"],
        [6, "Charge", "Movement", 18, 1.8, 75, 8, "Any", "Rush attack with momentum"],
        [7, "Parry", "Defense", 5, 0, 100, 2, "Melee", "Blocks and counters"],
        [8, "Berserker Rage", "Buff", 30, 2.5, 70, 15, "Any", "Increased damage, reduced defense"]
    ]
    
    for i, skill in enumerate(combat_skills, 2):
        for j, value in enumerate(skill, 1):
            ws_combat.cell(row=i, column=j, value=value)
    
    wb.save("skills.xlsx")
    print("Created skills.xlsx")

def create_monsters_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Monsters"
    
    headers = ["ID", "Name", "MonsterType", "Level", "HP", "MP", "Attack", "Defense", "Speed", "EXP_Reward", "Gold_Drop", "Special_Abilities", "Weakness"]
    for i, header in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=header)
    
    monsters = [
        [1, "Goblin", "Humanoid", 3, 45, 0, 15, 8, 25, 25, "5-15", "Sneak Attack", "Light"],
        [2, "Orc Warrior", "Humanoid", 8, 120, 10, 35, 20, 15, 80, "20-40", "Rage", "Magic"],
        [3, "Fire Drake", "Dragon", 15, 300, 50, 60, 35, 30, 200, "80-120", "Fire Breath", "Ice"],
        [4, "Skeleton", "Undead", 5, 60, 0, 20, 15, 10, 40, "10-25", "Bone Throw", "Holy"],
        [5, "Ice Elemental", "Elemental", 12, 180, 80, 40, 25, 20, 120, "50-80", "Ice Storm", "Fire"],
        [6, "Shadow Wolf", "Beast", 7, 90, 5, 30, 18, 40, 60, "15-30", "Shadow Step", "Light"],
        [7, "Ancient Lich", "Undead", 25, 500, 200, 80, 40, 15, 400, "200-300", "Death Magic", "Holy"],
        [8, "Stone Golem", "Construct", 18, 400, 0, 70, 60, 5, 250, "100-150", "Stone Skin", "Lightning"],
        [9, "Vampire Lord", "Undead", 22, 350, 100, 75, 30, 35, 300, "150-250", "Life Drain", "Holy"],
        [10, "Red Dragon", "Dragon", 30, 800, 150, 100, 50, 25, 500, "400-600", "Dragon Fire", "Ice"]
    ]
    
    for i, monster in enumerate(monsters, 2):
        for j, value in enumerate(monster, 1):
            ws.cell(row=i, column=j, value=value)
    
    wb.save("monsters.xlsx")
    print("Created monsters.xlsx")

def create_quests_excel():
    wb = Workbook()
    
    # Main Quests
    ws_main = wb.active
    ws_main.title = "Main_Quests"
    
    main_headers = ["ID", "Name", "Description", "Level_Required", "EXP_Reward", "Gold_Reward", "Item_Reward", "Prerequisites", "Location", "Quest_Giver"]
    for i, header in enumerate(main_headers, 1):
        ws_main.cell(row=1, column=i, value=header)
    
    main_quests = [
        [1, "The Hero's Journey", "Begin your adventure", 1, 100, 50, "Iron Sword", "None", "Starting Village", "Village Elder"],
        [2, "Goblin Menace", "Clear the goblin camp", 3, 250, 150, "Leather Armor", "Quest 1", "Forest Outskirts", "Guard Captain"],
        [3, "The Lost Artifact", "Find the ancient relic", 8, 500, 300, "Magic Ring", "Quest 2", "Ancient Ruins", "Wise Sage"],
        [4, "Dragon's Lair", "Defeat the fire drake", 15, 1000, 800, "Dragon Scale", "Quest 3", "Mountain Cave", "Knight Commander"],
        [5, "The Final Battle", "Confront the Dark Lord", 25, 2000, 1500, "Legendary Weapon", "Quest 4", "Dark Castle", "High Priestess"]
    ]
    
    for i, quest in enumerate(main_quests, 2):
        for j, value in enumerate(quest, 1):
            ws_main.cell(row=i, column=j, value=value)
    
    # Side Quests
    ws_side = wb.create_sheet("Side_Quests")
    side_headers = ["ID", "Name", "Description", "Level_Required", "EXP_Reward", "Gold_Reward", "Item_Reward", "QuestType", "Repeatable"]
    for i, header in enumerate(side_headers, 1):
        ws_side.cell(row=1, column=i, value=header)
    
    side_quests = [
        [1, "Herb Gathering", "Collect 10 healing herbs", 1, 50, 25, "Health Potion", "Collection", "Yes"],
        [2, "Merchant's Delivery", "Deliver package to next town", 2, 75, 40, "None", "Delivery", "Yes"],
        [3, "Wolf Hunt", "Kill 5 wolves", 5, 150, 80, "Wolf Pelt", "Hunting", "Yes"],
        [4, "Lost Cat", "Find the missing cat", 1, 30, 15, "Cat Treats", "Search", "No"],
        [5, "Bandit Camp", "Clear the bandit hideout", 10, 300, 200, "Bandit Armor", "Combat", "No"],
        [6, "Rare Minerals", "Mine 20 rare crystals", 8, 200, 120, "Crystal", "Mining", "Yes"],
        [7, "Ancient Tome", "Retrieve the lost spellbook", 12, 400, 250, "Spell Scroll", "Exploration", "No"],
        [8, "Tournament", "Win the fighting tournament", 15, 600, 500, "Champion's Ring", "Combat", "No"]
    ]
    
    for i, quest in enumerate(side_quests, 2):
        for j, value in enumerate(quest, 1):
            ws_side.cell(row=i, column=j, value=value)
    
    wb.save("quests.xlsx")
    print("Created quests.xlsx")

if __name__ == "__main__":
    create_characters_excel()
    create_items_excel()
    create_skills_excel()
    create_monsters_excel()
    create_quests_excel()
    print("All RPG Excel files created successfully!")