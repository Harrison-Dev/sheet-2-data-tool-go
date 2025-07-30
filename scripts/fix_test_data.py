#!/usr/bin/env python3
import openpyxl
import os

def fix_excel_files():
    test_data_dir = "test-data"
    
    for filename in os.listdir(test_data_dir):
        if filename.endswith('.xlsx'):
            filepath = os.path.join(test_data_dir, filename)
            print(f"Fixing {filename}...")
            
            wb = openpyxl.load_workbook(filepath)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Check if first column header is "ID" and change to "Id"
                if ws.cell(row=1, column=1).value == "ID":
                    ws.cell(row=1, column=1, value="Id")
                    print(f"  - Changed ID to Id in sheet {sheet_name}")
            
            wb.save(filepath)
            print(f"  ✓ {filename} updated")

if __name__ == "__main__":
    fix_excel_files()
    print("All files fixed!")